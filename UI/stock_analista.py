"""
stock_analista.py  –  Vista de Stock para Analista
===================================================
Ventana que muestra el inventario actual con columnas adaptadas
al sistema de sustancias primarias de CECIF, con exportación a Excel.
"""
import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk

from config.config import COLORS
from logica import inventario_mov_logica as inv
from logica import movimientos_common as common
from UI._mov_utils import (
    apply_default_window,
    attach_treeview_sorting,
    bind_horizontal_wheel_scroll,
    configure_row_stripes,
    draw_title,
    style_treeview,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def open_window(master):
    StockAnalistaWindow(master)


def _safe_float(value) -> float:
    try:
        return float(str(value).replace(",", ".").strip())
    except (TypeError, ValueError):
        return 0.0


COLUMNS = [
    ("codigo",            "Código",              90),
    ("nombre",            "Nombre",             200),
    ("propiedad",         "Propiedad",           110),
    ("lote",              "Lote",                 90),
    ("unidad",            "Unidad",               70),
    ("presentacion",      "Presentación",        100),
    ("stock",             "Stock",                80),
    ("cantidad_botellas", "Cant. Botellas",      170),
    ("ubicacion",         "Ubicación",            110),
    ("no_caja",           "No. Caja",             80),
    ("fecha_vencimiento", "F. Vencimiento",       110),
    ("fabricante",        "Fabricante",           160),
    ("potencia",          "Potencia",             100),
    ("catalogo",          "Catálogo",             110),
    ("condicion_alm",     "Cond. Almacenamiento", 180),
]


class StockAnalistaWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Sistema de Gestión - Stock Analista")
        self.configure(bg=COLORS["secondary"])
        apply_default_window(self, min_width=1100, min_height=680)

        self._all_rows: list[dict] = []
        self._filtered_rows: list[dict] = []
        self._tooltip = None
        self._current_page = 1
        self._page_size = 50
        self._total_pages = 1
        self._total_records = 0

        self.v_search = tk.StringVar()
        self.v_por_pagina = tk.StringVar(value="50")

        draw_title(self, "Sistema de Gestión – Stock Analista")
        self._build_ui()
        self._load_data()

    # ------------------------------------------------------------------
    # Construcción de UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        wrap = tk.Frame(self, bg=COLORS["secondary"])
        wrap.pack(fill="both", expand=True, padx=10, pady=10)

        # ── Fila de búsqueda ──────────────────────────────────────────
        top = tk.Frame(wrap, bg=COLORS["secondary"])
        top.pack(fill="x", pady=(0, 6))

        tk.Label(top, text="Buscar:", bg=COLORS["secondary"],
                 fg=COLORS["text_dark"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
        search_entry = tk.Entry(top, textvariable=self.v_search, width=30,
                                font=("Segoe UI", 9))
        search_entry.pack(side="left", padx=(0, 6))
        search_entry.bind("<Return>", lambda _: self._apply_search())

        self._btn(top, "Buscar", COLORS["primary"], self._apply_search).pack(side="left", padx=(0, 4))
        self._btn(top, "⟳ Actualizar", COLORS["success"], self._load_data).pack(side="left", padx=(0, 4))
        self._btn(top, "📥 Exportar Excel", "#1565C0", self._export_excel).pack(side="left", padx=(0, 4))
        self._btn(top, "📋 Histórico Lote", "#5C3B8E", self._show_lote_history).pack(side="left", padx=(0, 4))
        self._btn(top, "Salir", "#6C757D", self.destroy).pack(side="right")

        # ── Leyenda ───────────────────────────────────────────────────
        legend = tk.Frame(wrap, bg=COLORS["secondary"])
        legend.pack(fill="x", pady=(0, 4))
        tk.Label(legend, text="  VENCIDO  ", bg="#FFD6D6", fg="#8B0000",
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 4))
        tk.Label(legend, text="  PRÓXIMO A VENCER  ", bg="#FFF3CD", fg="#7A5D00",
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 4))
        tk.Label(legend, text="  AGOTADO  ", bg="#E0E0E0", fg="#424242",
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=(0, 4))

        # ── Tabla ─────────────────────────────────────────────────────
        table_frame = tk.Frame(wrap, bg=COLORS["secondary"])
        table_frame.pack(fill="both", expand=True)
        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        cols = [c[0] for c in COLUMNS]
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=18)

        style_treeview(self)

        sy = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        sx = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        for key, title, width in COLUMNS:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, anchor="center", minwidth=50, stretch=True)

        attach_treeview_sorting(self.tree)

        configure_row_stripes(self.tree)
        bind_horizontal_wheel_scroll(self.tree)
        self.tree.tag_configure("vencido", background="#FFD6D6")
        self.tree.tag_configure("proximo", background="#FFF3CD")
        self.tree.tag_configure("agotado", background="#E0E0E0")
        self.tree.tag_configure("ok",      background="#DFF7E2")

        self.tree.bind("<Motion>", self._on_tree_motion)
        self.tree.bind("<Leave>", lambda _e: self._hide_tooltip())

        # ── Paginación ────────────────────────────────────────────────
        pag_frame = tk.Frame(wrap, bg=COLORS["secondary"])
        pag_frame.pack(fill="x", pady=(4, 0))

        btn_s = dict(bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold"),
                     relief="flat", bd=0, padx=8, pady=3, cursor="hand2")
        self.btn_prev = tk.Button(pag_frame, text="< Anterior",
                                  command=self._prev_page, **btn_s)
        self.btn_prev.pack(side="left", padx=2)
        self.lbl_page = tk.Label(pag_frame, text="Página 1 de 1",
                                 bg=COLORS["secondary"], fg=COLORS["text_dark"],
                                 font=("Segoe UI", 9))
        self.lbl_page.pack(side="left", padx=10)
        self.btn_next = tk.Button(pag_frame, text="Siguiente >",
                                  command=self._next_page, **btn_s)
        self.btn_next.pack(side="left", padx=2)

        tk.Label(pag_frame, text="Mostrar:", bg=COLORS["secondary"],
                 fg=COLORS["text_dark"], font=("Segoe UI", 9)).pack(side="left", padx=(15, 4))
        self.combo_page_size = ttk.Combobox(
            pag_frame, values=["20", "50", "100", "200"],
            state="readonly", width=6, font=("Segoe UI", 9),
        )
        self.combo_page_size.set("50")
        self.combo_page_size.bind("<<ComboboxSelected>>", self._on_page_size_change)
        self.combo_page_size.pack(side="left", padx=2)

        self.lbl_total = tk.Label(pag_frame, text="", bg=COLORS["secondary"],
                                  fg=COLORS["text_muted"], font=("Segoe UI", 9, "italic"))
        self.lbl_total.pack(side="left", padx=(12, 0))

    # ------------------------------------------------------------------
    # Carga de datos
    # ------------------------------------------------------------------

    def _load_data(self):
        self.v_search.set("")
        snapshot = inv.cargar_snapshot()
        self._all_rows = []
        for row in snapshot:
            if str(row.get("estado", "ACTIVA")).upper() in ("ANULADA", "ANULADO"):
                continue
            stock_val = _safe_float(row.get("cantidad_actual", row.get("stock", 0)))
            row["cantidad_botellas"] = common.texto_envases(stock_val)
            self._all_rows.append(row)
        self._filtered_rows = list(self._all_rows)
        self._current_page = 1
        self._total_records = len(self._filtered_rows)
        self._total_pages = max(1, -(-self._total_records // self._page_size))
        self._render_page()

    def _apply_search(self):
        query = self.v_search.get().strip().lower()
        if not query:
            self._filtered_rows = list(self._all_rows)
        else:
            self._filtered_rows = [
                r for r in self._all_rows
                if query in str(r.get("codigo", "")).lower()
                or query in str(r.get("nombre", "")).lower()
                or query in str(r.get("lote", "")).lower()
                or query in str(r.get("ubicacion", "")).lower()
                or query in str(r.get("fabricante", "")).lower()
            ]
        self._current_page = 1
        self._total_records = len(self._filtered_rows)
        self._total_pages = max(1, -(-self._total_records // self._page_size))
        self._render_page()

    def _render_page(self):
        self.tree.delete(*self.tree.get_children())
        start = (self._current_page - 1) * self._page_size
        page_rows = self._filtered_rows[start:start + self._page_size]

        for idx, row in enumerate(page_rows):
            alarma_fv = str(row.get("alarma_fv", "")).upper()
            alarma_stock = str(row.get("alarma_stock", "")).upper()
            stock_val = _safe_float(row.get("cantidad_actual", row.get("stock", 0)))

            if alarma_fv == "VENCIDO":
                tag = "vencido"
            elif alarma_fv == "PROXIMO A VENCER":
                tag = "proximo"
            elif stock_val <= 0:
                tag = "agotado"
            elif alarma_fv == "OK":
                tag = "ok"
            else:
                tag = "par" if idx % 2 == 0 else "impar"

            values = tuple(
                row.get(key, "") for key, _, _ in COLUMNS
            )
            self.tree.insert("", "end", iid=str(row.get("id", "")), values=values, tags=(tag,))

        self._update_pagination()

    # ------------------------------------------------------------------
    # Paginación
    # ------------------------------------------------------------------

    def _prev_page(self):
        if self._current_page > 1:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_page < self._total_pages:
            self._current_page += 1
            self._render_page()

    def _on_page_size_change(self, _event=None):
        try:
            self._page_size = int(self.combo_page_size.get())
            self._current_page = 1
            self._total_pages = max(1, -(-self._total_records // self._page_size))
            self._render_page()
        except ValueError:
            pass

    def _update_pagination(self):
        self.btn_prev.config(state=tk.NORMAL if self._current_page > 1 else tk.DISABLED)
        self.btn_next.config(state=tk.NORMAL if self._current_page < self._total_pages else tk.DISABLED)
        self.lbl_page.config(text=f"Página {self._current_page} de {max(self._total_pages, 1)}")
        self.lbl_total.config(text=f"({self._total_records} registros)")

    # ------------------------------------------------------------------
    # Exportación Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        if not _OPENPYXL:
            messagebox.showerror(
                "Dependencia faltante",
                "openpyxl no está instalado.\nInstala con: pip install openpyxl",
                parent=self,
            )
            return

        rows = self._filtered_rows
        if not rows:
            messagebox.showwarning("Sin datos", "No hay datos para exportar.", parent=self)
            return

        default_name = f"stock_analista_{date.today().strftime('%Y%m%d')}.xlsx"
        filepath = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar Stock Analista",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if not filepath:
            return

        try:
            self._write_excel(filepath, rows)
            messagebox.showinfo("Exportado",
                                f"Reporte guardado exitosamente:\n{filepath}", parent=self)
        except PermissionError:
            messagebox.showerror(
                "Error",
                "No se pudo guardar el archivo. Verifica que no esté abierto en Excel.",
                parent=self,
            )
        except Exception as exc:
            messagebox.showerror("Error", f"Error al exportar:\n{exc}", parent=self)

    @staticmethod
    def _write_excel(filepath: str, rows: list[dict]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Analista"

        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="0F4C97")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        vencido_fill = PatternFill("solid", fgColor="FFD6D6")
        proximo_fill = PatternFill("solid", fgColor="FFF3CD")
        agotado_fill = PatternFill("solid", fgColor="E0E0E0")
        ok_fill      = PatternFill("solid", fgColor="DFF7E2")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = [title for _, title, _ in COLUMNS]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        col_widths = {key: max(len(title), 8) for key, title, _ in COLUMNS}

        for row_idx, row in enumerate(rows, start=2):
            stock_val  = _safe_float(row.get("cantidad_actual", row.get("stock", 0)))
            alarma_fv  = str(row.get("alarma_fv", "")).upper()

            if alarma_fv == "VENCIDO":
                row_fill = vencido_fill
            elif alarma_fv == "PROXIMO A VENCER":
                row_fill = proximo_fill
            elif stock_val <= 0:
                row_fill = agotado_fill
            elif alarma_fv == "OK":
                row_fill = ok_fill
            else:
                row_fill = None

            for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
                val = row.get(key, "")
                if val is None:
                    val = ""
                elif isinstance(val, bool):
                    val = "Sí" if val else "No"
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin_border
                if row_fill:
                    cell.fill = row_fill
                # Calcular ancho
                str_val = str(val)
                if len(str_val) > col_widths.get(key, 0):
                    col_widths[key] = len(str_val)

        # Ajustar anchos de columna
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(col_widths[key] + 4, 50)

        ws.row_dimensions[1].height = 32
        ws.freeze_panes = "A2"
        wb.save(filepath)

    # ------------------------------------------------------------------
    # Tooltip condicion_alm
    # ------------------------------------------------------------------

    def _on_tree_motion(self, event):
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if not row_id or not col_id:
            self._hide_tooltip()
            return

        col_idx = int(col_id.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(COLUMNS):
            self._hide_tooltip()
            return

        key = COLUMNS[col_idx][0]
        if key != "condicion_alm":
            self._hide_tooltip()
            return

        values = self.tree.item(row_id, "values")
        if col_idx >= len(values):
            self._hide_tooltip()
            return

        text = str(values[col_idx])
        if len(text) < 20:
            self._hide_tooltip()
            return

        self._show_tooltip(event.x_root + 12, event.y_root + 12, text)

    def _show_tooltip(self, x, y, text):
        if self._tooltip is not None:
            lbl = getattr(self._tooltip, "_lbl", None)
            if lbl is not None and lbl.cget("text") == text:
                self._tooltip.geometry(f"+{x}+{y}")
                return
            self._hide_tooltip()

        tip = tk.Toplevel(self)
        tip.wm_overrideredirect(True)
        tip.attributes("-topmost", True)
        tip.geometry(f"+{x}+{y}")
        lbl = tk.Label(
            tip,
            text=text,
            justify="left",
            bg="#FFFFE0",
            fg="#202020",
            relief="solid",
            bd=1,
            padx=6,
            pady=4,
            wraplength=520,
            font=("Segoe UI", 9),
        )
        lbl.pack()
        tip._lbl = lbl
        self._tooltip = tip

    def _hide_tooltip(self):
        if self._tooltip is not None:
            self._tooltip.destroy()
            self._tooltip = None

    # ------------------------------------------------------------------
    # Histórico Lote
    # ------------------------------------------------------------------

    def _show_lote_history(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo(
                "Selección requerida",
                "Selecciona un ítem del inventario para ver su historial de movimientos.",
                parent=self,
            )
            return

        iid = selected[0]
        inv_row = next(
            (r for r in self._filtered_rows if str(r.get("id", "")) == iid), None
        )
        if inv_row is None:
            messagebox.showerror("Error", "No se encontró el registro.", parent=self)
            return

        id_inventario = inv_row.get("id")
        codigo    = inv_row.get("codigo", "")
        nombre    = inv_row.get("nombre", "")
        lote      = inv_row.get("lote", "")
        unidad    = inv_row.get("unidad", "")
        cant_act  = _safe_float(inv_row.get("cantidad_actual", inv_row.get("stock", 0)))

        # Cargar y filtrar salidas de este item
        all_salidas = common.cargar_salidas()
        salidas_item = [
            s for s in all_salidas
            if str(s.get("id_inventario", s.get("id_entrada", ""))) == str(id_inventario)
        ]
        salidas_item.sort(key=lambda x: str(x.get("fecha_hora", "")))

        # Sumar sólo salidas activas para reconstruir total entrada
        total_sal_activas = sum(
            _safe_float(s.get("cantidad", 0))
            for s in salidas_item
            if str(s.get("estado", "ACTIVA")).upper() not in ("ANULADA", "ANULADO")
        )
        total_entrada = cant_act + total_sal_activas

        # Calcular stock remanente cronológicamente
        stock_acum = total_entrada
        rows_hist = []
        for s in salidas_item:
            qty = _safe_float(s.get("cantidad", 0))
            anulada = str(s.get("estado", "ACTIVA")).upper() in ("ANULADA", "ANULADO")
            if not anulada:
                stock_acum -= qty
            stock_rem = round(stock_acum, 4)
            cant_bot  = common.texto_envases(stock_rem)
            rows_hist.append({
                "fecha_hora":       s.get("fecha_hora", ""),
                "usuario":          s.get("usuario_nombre", ""),
                "tipo_salida":      s.get("tipo_salida", ""),
                "cantidad":         qty,
                "unidad":           unidad,
                "stock_rem":        stock_rem,
                "cant_botellas":    cant_bot,
                "estado":           s.get("estado", "ACTIVA"),
                "observacion":      s.get("observacion", ""),
            })
        rows_hist.reverse()   # más reciente primero

        # ── Ventana popup ─────────────────────────────────────────────
        win = tk.Toplevel(self)
        win.title(f"Histórico de Salidas  –  {codigo} / Lote {lote}")
        win.configure(bg=COLORS["secondary"])
        apply_default_window(win, min_width=1050, min_height=420)

        tk.Label(
            win,
            text=f"Producto: {nombre}   |   Código: {codigo}   |   Lote: {lote}   |   Stock actual: {cant_act} {unidad}",
            bg=COLORS["primary"], fg="white",
            font=("Segoe UI", 10, "bold"), pady=7,
        ).pack(fill="x")

        table_wrap = tk.Frame(win, bg=COLORS["secondary"])
        table_wrap.pack(fill="both", expand=True, padx=10, pady=10)
        table_wrap.rowconfigure(0, weight=1)
        table_wrap.columnconfigure(0, weight=1)

        hist_cols = [
            ("fecha_hora",    "Fecha/Hora",              145),
            ("usuario",       "Usuario",                  130),
            ("tipo_salida",   "Tipo Salida",              130),
            ("cantidad",      "Cantidad Salida",          100),
            ("unidad",        "Unidad",                    80),
            ("stock_rem",     "Stock Remanente",          120),
            ("cant_botellas", "Cant. Botellas Restantes", 200),
            ("estado",        "Estado",                    90),
            ("observacion",   "Observación",              240),
        ]
        col_keys = [c[0] for c in hist_cols]

        style_treeview(win, style_name="Hist.Treeview")

        tree = ttk.Treeview(
            table_wrap, columns=col_keys, show="headings",
            height=14, style="Hist.Treeview",
        )
        sy = ttk.Scrollbar(table_wrap, orient="vertical",   command=tree.yview)
        sx = ttk.Scrollbar(table_wrap, orient="horizontal",  command=tree.xview)
        tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        for key, title, width in hist_cols:
            tree.heading(key, text=title)
            tree.column(key, width=width, anchor="center", minwidth=50, stretch=True)

        configure_row_stripes(tree)
        bind_horizontal_wheel_scroll(tree)
        tree.tag_configure("anulada", background="#FADADD", foreground="#666")

        if rows_hist:
            for idx, h in enumerate(rows_hist):
                tag = "anulada" if h["estado"].upper() in ("ANULADA", "ANULADO") else ("par" if idx % 2 == 0 else "impar")
                tree.insert("", "end",
                             values=tuple(h[k] for k in col_keys),
                             tags=(tag,))
        else:
            tree.insert("", "end", values=("Sin movimientos registrados",) + ("",) * (len(col_keys) - 1))

        bottom = tk.Frame(win, bg=COLORS["secondary"])
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        tk.Label(
            bottom,
            text=f"Total entradas: {round(total_entrada, 4)} {unidad}   |   Total salidas activas: {round(total_sal_activas, 4)} {unidad}",
            bg=COLORS["secondary"], fg=COLORS["text_muted"],
            font=("Segoe UI", 9, "italic"),
        ).pack(side="left")
        tk.Button(
            bottom, text="Cerrar", bg="#6C757D", fg="white",
            font=("Segoe UI", 9, "bold"), relief="flat",
            bd=0, padx=14, pady=5, cursor="hand2",
            command=win.destroy,
        ).pack(side="right")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _btn(self, parent, text, bg, cmd):
        return tk.Button(
            parent, text=text, bg=bg, fg="white",
            font=("Segoe UI", 9, "bold"), relief="flat",
            bd=0, padx=10, pady=5, cursor="hand2", command=cmd,
        )
