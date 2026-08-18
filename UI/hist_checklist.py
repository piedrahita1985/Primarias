"""
hist_checklist.py  –  Historial de Listas de Chequeo
=====================================================
Ventana que muestra el historial de listas de chequeo (CECIF y Cliente),
con filtros por tipo, código de sustancia, lote y fechas, y exportación a Excel.
"""
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

from config.config import COLORS
from logica import check_logica as chk
from logica import movimientos_common as common
from UI._mov_utils import (
    apply_default_window,
    bind_horizontal_wheel_scroll,
    configure_row_stripes,
    draw_title,
    get_date_value,
    make_date_widget,
    style_treeview,
)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def open_window(master):
    HistCheckListWindow(master)


# ── Campos de verificación CECIF ────────────────────────────────────────────
CECIF_CHECK_FIELDS = [
    ("ver_nombre",                  "Nombre"),
    ("ver_no_lote",                 "No. de Lote"),
    ("ver_cantidad",                "Cantidad"),
    ("ver_rotulo_identificacion",   "Rótulo de Identificación"),
    ("ver_fecha_fabricacion",       "Fecha de Fabricación"),
    ("ver_fecha_vencimiento",       "Fecha de Vencimiento"),
    ("ver_fabricante",              "Fabricante"),
    ("ver_rotulos_seguridad",       "Rótulos de seguridad, sellos y precintos"),
    ("ver_ficha_seguridad",         "Ficha de Seguridad"),
    ("ver_certificado_calidad",     "Certificado de Calidad"),
    ("ver_golpes_roturas",          "Se evidencian Golpes, Roturas u Otros"),
    ("ver_cumple_especificaciones", "Cumple con las especificaciones requeridas"),
]

# ── Campos de verificación Cliente – nuevas ─────────────────────────────────
CLIENTE_NUEVAS_FIELDS = [
    ("vn_nombre",                  "NVA: Nombre"),
    ("vn_no_lote",                 "NVA: No. de Lote"),
    ("vn_cantidad",                "NVA: Cantidad"),
    ("vn_rotulo_identificacion",   "NVA: Rótulo de Identificación"),
    ("vn_fecha_fabricacion",       "NVA: Fecha de Fabricación"),
    ("vn_fecha_vencimiento",       "NVA: Fecha de Vencimiento"),
    ("vn_fabricante",              "NVA: Fabricante"),
    ("vn_rotulos_seguridad",       "NVA: Rótulos de seguridad"),
    ("vn_ficha_seguridad",         "NVA: Ficha de Seguridad"),
    ("vn_certificado_calidad",     "NVA: Certificado de Calidad"),
    ("vn_golpes_roturas",          "NVA: Golpes/Roturas"),
    ("vn_cumple_especificaciones", "NVA: Cumple especificaciones"),
]

# ── Campos de verificación Cliente – destapadas ─────────────────────────────
CLIENTE_DEST_FIELDS = [
    ("vd_nombre",                        "DEST: Nombre"),
    ("vd_no_lote",                       "DEST: No. de Lote"),
    ("vd_rotulo_identificacion",         "DEST: Rótulo de Identificación"),
    ("vd_fecha_vencimiento",             "DEST: Fecha de Vencimiento"),
    ("vd_certificado_calidad",           "DEST: Certificado de Calidad"),
    ("vd_ficha_seguridad",               "DEST: Ficha de Seguridad"),
    ("vd_golpes_roturas",                "DEST: Golpes/Roturas"),
    ("vd_condiciones_almacenamiento",    "DEST: Condiciones almacenamiento"),
    ("vd_carta_correo",                  "DEST: Carta/correo"),
]


class HistCheckListWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Historial de Listas de Chequeo")
        self.configure(bg=COLORS["secondary"])
        apply_default_window(self, min_width=1100, min_height=680)

        # Catálogos
        maestras = common.cargar_maestras()
        all_sustancias = _load_all_sustancias()
        self._sustancias_by_id = common.map_by_id(all_sustancias)
        self._fabricantes_by_id = common.map_by_id(maestras["fabricantes"])
        self._usuarios_by_id: dict[int, dict] = {}
        _load_usuarios_into(self._usuarios_by_id)

        # Estado de filtros
        self.v_tipo = tk.StringVar(value="CECIF")
        self.v_codigo = tk.StringVar()
        self.v_lote = tk.StringVar()
        self._desde_habilitada = False
        self._hasta_habilitada = False

        # Paginación
        self._all_rows: list[dict] = []
        self._current_page = 1
        self._page_size = 50
        self._total_pages = 1
        self._total_records = 0

        draw_title(self, "Sistema de Gestión – Historial de Listas de Chequeo")
        self._build_ui()
        self.bind("<Escape>", lambda _: self.destroy())
        self._load_table()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _build_ui(self):
        wrap = tk.Frame(self, bg=COLORS["secondary"])
        wrap.pack(fill="both", expand=True, padx=10, pady=10)

        # ── Filtros ───────────────────────────────────────────────────
        filter_row = tk.Frame(wrap, bg=COLORS["secondary"])
        filter_row.pack(fill="x", pady=(0, 6))

        # Tipo (CECIF / Cliente)
        tk.Label(filter_row, text="Tipo:", bg=COLORS["secondary"],
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        tipo_combo = ttk.Combobox(
            filter_row, textvariable=self.v_tipo,
            values=["CECIF", "Cliente"], state="readonly", width=10,
        )
        tipo_combo.pack(side="left", padx=(0, 14))
        tipo_combo.bind("<<ComboboxSelected>>", lambda _: self._on_tipo_change())

        # Código sustancia
        tk.Label(filter_row, text="Código:", bg=COLORS["secondary"],
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        tk.Entry(filter_row, textvariable=self.v_codigo, width=14,
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 12))
        self.v_codigo.trace_add("write", lambda *_: self._load_table())

        # Lote
        tk.Label(filter_row, text="Lote:", bg=COLORS["secondary"],
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        tk.Entry(filter_row, textvariable=self.v_lote, width=14,
                 font=("Segoe UI", 9)).pack(side="left", padx=(0, 12))
        self.v_lote.trace_add("write", lambda *_: self._load_table())

        # Fecha Desde
        tk.Label(filter_row, text="Desde:", bg=COLORS["secondary"],
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        self.w_desde = make_date_widget(filter_row)
        self.w_desde.pack(side="left", padx=(0, 12))
        self.w_desde.bind("<<DateEntrySelected>>", lambda _: self._toggle_date("desde", True))

        # Fecha Hasta
        tk.Label(filter_row, text="Hasta:", bg=COLORS["secondary"],
                 font=("Segoe UI", 10)).pack(side="left", padx=(0, 4))
        self.w_hasta = make_date_widget(filter_row)
        self.w_hasta.pack(side="left", padx=(0, 10))
        self.w_hasta.bind("<<DateEntrySelected>>", lambda _: self._toggle_date("hasta", True))

        self._btn(filter_row, "Filtrar", COLORS["primary"],
                  self._load_table).pack(side="left", padx=(0, 4))
        self._btn(filter_row, "⟳ Limpiar", "#B0B0B0",
                  self._clear_filters).pack(side="left", padx=(0, 4))
        self._btn(filter_row, "📥 Exportar Excel", "#1565C0",
                  self._export_excel).pack(side="right")

        # ── Tabla (reconstituida según tipo) ─────────────────────────
        self._table_wrap = tk.Frame(wrap, bg=COLORS["secondary"])
        self._table_wrap.pack(fill="both", expand=True)
        self.tree: ttk.Treeview | None = None
        self._build_tree()

        # ── Paginación ────────────────────────────────────────────────
        pag_frame = tk.Frame(wrap, bg=COLORS["secondary"])
        pag_frame.pack(fill="x", pady=(4, 0))

        btn_s = dict(bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold"),
                     relief="flat", bd=0, padx=8, pady=3, cursor="hand2")
        self.btn_prev = tk.Button(pag_frame, text="< Anterior",
                                  command=self._prev_page, **btn_s)
        self.btn_prev.pack(side="left", padx=2)
        self.lbl_page = tk.Label(pag_frame, text="Página 1 de 1",
                                 bg=COLORS["secondary"], fg=COLORS["text_dark"],
                                 font=("Segoe UI", 9))
        self.lbl_page.pack(side="left", padx=10)
        self.btn_next = tk.Button(pag_frame, text="Siguiente >",
                                  command=self._next_page, **btn_s)
        self.btn_next.pack(side="left", padx=2)

        tk.Label(pag_frame, text="Mostrar:", bg=COLORS["secondary"],
                 fg=COLORS["text_dark"], font=("Segoe UI", 9)).pack(side="left", padx=(14, 4))
        self.combo_page_size = ttk.Combobox(
            pag_frame, values=["20", "50", "100", "200"],
            state="readonly", width=6, font=("Segoe UI", 9),
        )
        self.combo_page_size.set("50")
        self.combo_page_size.bind("<<ComboboxSelected>>", self._on_page_size_change)
        self.combo_page_size.pack(side="left", padx=2)

        self.lbl_total = tk.Label(pag_frame, text="", bg=COLORS["secondary"],
                                  fg=COLORS["text_muted"], font=("Segoe UI", 9, "italic"))
        self.lbl_total.pack(side="left", padx=(12, 0))

        # ── Botón Salir ───────────────────────────────────────────────
        bottom = tk.Frame(self, bg=COLORS["secondary"])
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        self._btn(bottom, "Salir", "#6C757D", self.destroy).pack(side="right")

    def _build_tree(self):
        """Construye (o reconstruye) el Treeview según el tipo seleccionado."""
        for w in self._table_wrap.winfo_children():
            w.destroy()

        tipo = self.v_tipo.get()
        if tipo == "CECIF":
            cols_def = [
                ("id",            "ID",               50),
                ("fecha_hora",    "Fecha/Hora",       130),
                ("fecha_rec",     "Fecha Recepción",  110),
                ("proveedor",     "Proveedor/Fabric.", 160),
                ("orden_compra",  "Orden Compra",     110),
                ("codigo",        "Código",            80),
                ("nombre",        "Nombre Sustancia", 200),
                ("lote",          "Lote",              90),
                ("cantidad",      "Cantidad",          80),
                ("aprobo",        "Aprobó",           120),
                ("verifico",      "Verificó",         120),
                ("observaciones", "Observaciones",    180),
            ]
        else:
            cols_def = [
                ("id",            "ID",                50),
                ("fecha_hora",    "Fecha/Hora",       130),
                ("fecha_rec",     "Fecha Recepción",  110),
                ("nombre_cliente","Cliente",           140),
                ("codigo",        "Código",             80),
                ("nombre",        "Nombre Sustancia",  200),
                ("lote",          "Lote",               90),
                ("cantidad",      "Cantidad",           80),
                ("reviso",        "Revisó",            120),
                ("verifico",      "Verificó",          120),
                ("observaciones", "Observaciones",     180),
            ]

        self._cols_def = cols_def
        cols = [c[0] for c in cols_def]

        self._table_wrap.grid_rowconfigure(0, weight=1)
        self._table_wrap.grid_columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(
            self._table_wrap, columns=cols, show="headings", height=16,
        )
        style_treeview(self)

        sy = ttk.Scrollbar(self._table_wrap, orient="vertical", command=self.tree.yview)
        sx = ttk.Scrollbar(self._table_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        sy.grid(row=0, column=1, sticky="ns")
        sx.grid(row=1, column=0, sticky="ew")

        for key, title, width in cols_def:
            self.tree.heading(key, text=title,
                              command=lambda c=key: self._sort_column(c, False))
            self.tree.column(key, width=width, anchor="w", minwidth=40, stretch=True)

        configure_row_stripes(self.tree)
        bind_horizontal_wheel_scroll(self.tree)

    # ------------------------------------------------------------------
    # Carga de datos
    # ------------------------------------------------------------------

    def _on_tipo_change(self):
        self._build_tree()
        self._load_table()

    def _load_table(self):
        if self.tree is None:
            return

        tipo = self.v_tipo.get()
        codigo_f = self.v_codigo.get().strip().lower()
        lote_f   = self.v_lote.get().strip().lower()
        desde_f  = get_date_value(self.w_desde) if self._desde_habilitada else ""
        hasta_f  = get_date_value(self.w_hasta) if self._hasta_habilitada else ""

        if tipo == "CECIF":
            raw_rows = chk.cargar_cecif()
        else:
            raw_rows = chk.cargar_clientes()

        filtered = []
        for rec in raw_rows:
            # Filtro código sustancia
            sust_id = rec.get("id_sustancia")
            sust = self._sustancias_by_id.get(sust_id) if sust_id else None
            codigo = str(sust.get("codigo", "") if sust else "").strip().lower()

            if codigo_f and codigo_f not in codigo:
                continue

            # Filtro lote
            lote = str(rec.get("lote", "")).strip().lower()
            if lote_f and lote_f not in lote:
                continue

            # Filtro fechas
            fecha_str = str(rec.get("fecha_recepcion") or rec.get("fecha_hora", ""))[:10]
            if desde_f and fecha_str and fecha_str < desde_f:
                continue
            if hasta_f and fecha_str and fecha_str > hasta_f:
                continue

            filtered.append(rec)

        self._all_rows = filtered
        self._current_page = 1
        self._total_records = len(filtered)
        self._total_pages = max(1, -(-self._total_records // self._page_size))
        self._render_page()

    def _render_page(self):
        if self.tree is None:
            return
        self.tree.delete(*self.tree.get_children())
        start = (self._current_page - 1) * self._page_size
        page = self._all_rows[start:start + self._page_size]

        tipo = self.v_tipo.get()
        for idx, rec in enumerate(page):
            row_values = self._build_row_values(rec, tipo)
            tag = "par" if idx % 2 == 0 else "impar"
            self.tree.insert("", "end", values=row_values, tags=(tag,))

        self._update_pagination()

    def _build_row_values(self, rec: dict, tipo: str) -> tuple:
        sust_id = rec.get("id_sustancia")
        sust    = self._sustancias_by_id.get(sust_id) if sust_id else None
        codigo  = str(sust.get("codigo", "") if sust else "")
        nombre  = str(sust.get("nombre", "") if sust else "")
        lote    = str(rec.get("lote", ""))

        if tipo == "CECIF":
            fab_id    = rec.get("id_proveedor")
            fab       = self._fabricantes_by_id.get(fab_id) if fab_id else None
            proveedor = str(fab.get("fabricante", "") if fab else "")

            uid_apr   = rec.get("id_usuario_aprobo")
            uid_ver   = rec.get("id_usuario_verifico")
            aprobo    = self._nombre_usuario(uid_apr)
            verifico  = self._nombre_usuario(uid_ver)

            return (
                rec.get("id", ""),
                rec.get("fecha_hora", ""),
                rec.get("fecha_recepcion", ""),
                proveedor,
                rec.get("no_orden_compra", ""),
                codigo, nombre, lote,
                rec.get("cantidad", ""),
                aprobo, verifico,
                rec.get("observaciones", ""),
            )
        else:
            uid_rev  = rec.get("id_usuario_reviso")
            uid_ver  = rec.get("id_usuario_verifico")
            reviso   = self._nombre_usuario(uid_rev)
            verifico = self._nombre_usuario(uid_ver)

            return (
                rec.get("id", ""),
                rec.get("fecha_hora", ""),
                rec.get("fecha_recepcion", ""),
                rec.get("nombre_cliente", ""),
                codigo, nombre, lote,
                rec.get("cantidad", ""),
                reviso, verifico,
                rec.get("observaciones", ""),
            )

    def _nombre_usuario(self, uid) -> str:
        if uid is None:
            return ""
        u = self._usuarios_by_id.get(int(uid))
        return str(u.get("nombre", "") if u else "")

    # ------------------------------------------------------------------
    # Paginación
    # ------------------------------------------------------------------

    def _prev_page(self):
        if self._current_page > 1:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_page < self._total_pages:
            self._current_page += 1
            self._render_page()

    def _on_page_size_change(self, _event=None):
        try:
            self._page_size = int(self.combo_page_size.get())
            self._current_page = 1
            self._total_pages = max(1, -(-self._total_records // self._page_size))
            self._render_page()
        except ValueError:
            pass

    def _update_pagination(self):
        self.btn_prev.config(state=tk.NORMAL if self._current_page > 1 else tk.DISABLED)
        self.btn_next.config(state=tk.NORMAL if self._current_page < self._total_pages else tk.DISABLED)
        self.lbl_page.config(text=f"Página {self._current_page} de {max(self._total_pages, 1)}")
        self.lbl_total.config(text=f"({self._total_records} registros)")

    # ------------------------------------------------------------------
    # Ordenamiento
    # ------------------------------------------------------------------

    def _sort_column(self, col: str, reverse: bool):
        if self.tree is None:
            return
        items = [(self.tree.set(iid, col), iid) for iid in self.tree.get_children("")]
        items.sort(key=lambda x: x[0].lower(), reverse=reverse)
        for i, (_, iid) in enumerate(items):
            self.tree.move(iid, "", i)
        self.tree.heading(col, command=lambda: self._sort_column(col, not reverse))

    # ------------------------------------------------------------------
    # Filtros
    # ------------------------------------------------------------------

    def _toggle_date(self, which: str, enabled: bool):
        if which == "desde":
            self._desde_habilitada = enabled
        else:
            self._hasta_habilitada = enabled

    def _clear_date(self, which: str):
        self._toggle_date(which, False)
        widget = self.w_desde if which == "desde" else self.w_hasta
        var = getattr(widget, "_fallback_var", None)
        if var is not None:
            var.set("")
            return
        try:
            widget.delete(0, "end")
        except Exception:
            pass
        self._load_table()

    def _clear_filters(self):
        self.v_codigo.set("")
        self.v_lote.set("")
        self._clear_date("desde")
        self._clear_date("hasta")
        self._load_table()

    # ------------------------------------------------------------------
    # Exportación Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        if not _OPENPYXL:
            messagebox.showerror(
                "Dependencia faltante",
                "openpyxl no está instalado.\nInstala con: pip install openpyxl",
                parent=self,
            )
            return

        rows = self._all_rows
        if not rows:
            messagebox.showwarning("Sin datos",
                                   "No hay registros para exportar con los filtros actuales.",
                                   parent=self)
            return

        tipo = self.v_tipo.get()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"HistCheckList_{tipo}_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(
            parent=self,
            title=f"Guardar Historial de Checklists – {tipo}",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if not filepath:
            return

        try:
            if tipo == "CECIF":
                self._write_excel_cecif(filepath, rows)
            else:
                self._write_excel_cliente(filepath, rows)
            messagebox.showinfo("Exportado",
                                f"Reporte guardado exitosamente:\n{filepath}", parent=self)
        except PermissionError:
            messagebox.showerror(
                "Error",
                "No se pudo guardar el archivo. Verifica que no esté abierto en Excel.",
                parent=self,
            )
        except Exception as exc:
            messagebox.showerror("Error", f"Error al exportar:\n{exc}", parent=self)

    # ── Excel CECIF ───────────────────────────────────────────────────

    def _write_excel_cecif(self, filepath: str, rows: list[dict]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "CheckList CECIF"

        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="0F4C97")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        si_fill      = PatternFill("solid", fgColor="DFF3E3")
        no_fill      = PatternFill("solid", fgColor="FADBD8")
        na_fill      = PatternFill("solid", fgColor="FFF9C4")
        thin         = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        main_headers = [
            "ID", "Fecha/Hora", "Fecha Recepción", "Proveedor/Fabricante",
            "Orden Compra", "Código", "Nombre Sustancia", "Lote",
            "Cantidad", "Aprobó", "Verificó", "Observaciones",
        ]
        check_headers = [title for _, title in CECIF_CHECK_FIELDS]
        all_headers = main_headers + check_headers

        for col_idx, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        main_widths = [6, 18, 16, 22, 14, 12, 30, 14, 10, 20, 20, 30]

        for row_idx, rec in enumerate(rows, start=2):
            sust_id = rec.get("id_sustancia")
            sust    = self._sustancias_by_id.get(sust_id) if sust_id else None
            fab_id  = rec.get("id_proveedor")
            fab     = self._fabricantes_by_id.get(fab_id) if fab_id else None

            main_values = [
                rec.get("id", ""),
                rec.get("fecha_hora", ""),
                rec.get("fecha_recepcion", ""),
                str(fab.get("fabricante", "") if fab else ""),
                rec.get("no_orden_compra", ""),
                str(sust.get("codigo", "") if sust else ""),
                str(sust.get("nombre", "") if sust else ""),
                rec.get("lote", ""),
                rec.get("cantidad", ""),
                self._nombre_usuario(rec.get("id_usuario_aprobo")),
                self._nombre_usuario(rec.get("id_usuario_verifico")),
                rec.get("observaciones", ""),
            ]

            alt = PatternFill("solid", fgColor="F0F4FC") if row_idx % 2 == 0 else None

            for col_idx, val in enumerate(main_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin
                if alt:
                    cell.fill = alt

            for ci, (field, _) in enumerate(CECIF_CHECK_FIELDS):
                resp = str(rec.get(field, "") or "").strip().upper()
                cell = ws.cell(row=row_idx, column=len(main_headers) + ci + 1, value=resp)
                cell.alignment = center_align
                cell.border = thin
                if resp == "SI":
                    cell.fill = si_fill
                elif resp == "NO":
                    cell.fill = no_fill
                elif resp in ("NA", "N/A", "NONE", ""):
                    cell.fill = na_fill

        # Anchos
        for col_idx, w in enumerate(main_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        for ci in range(len(CECIF_CHECK_FIELDS)):
            ws.column_dimensions[
                ws.cell(row=1, column=len(main_headers) + ci + 1).column_letter
            ].width = 20

        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"
        wb.save(filepath)

    # ── Excel Cliente ─────────────────────────────────────────────────

    def _write_excel_cliente(self, filepath: str, rows: list[dict]) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "CheckList Cliente"

        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="2E7D32")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        si_fill      = PatternFill("solid", fgColor="DFF3E3")
        no_fill      = PatternFill("solid", fgColor="FADBD8")
        na_fill      = PatternFill("solid", fgColor="FFF9C4")
        thin         = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        main_headers = [
            "ID", "Fecha/Hora", "Fecha Recepción", "Nombre Cliente",
            "Código", "Nombre Sustancia", "Lote", "Cantidad",
            "Revisó", "Verificó", "Observaciones",
        ]
        check_nva_headers  = [title for _, title in CLIENTE_NUEVAS_FIELDS]
        check_dest_headers = [title for _, title in CLIENTE_DEST_FIELDS]
        all_headers = main_headers + check_nva_headers + check_dest_headers

        for col_idx, h in enumerate(all_headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        main_widths = [6, 18, 16, 22, 12, 30, 14, 10, 20, 20, 30]

        for row_idx, rec in enumerate(rows, start=2):
            sust_id = rec.get("id_sustancia")
            sust    = self._sustancias_by_id.get(sust_id) if sust_id else None

            main_values = [
                rec.get("id", ""),
                rec.get("fecha_hora", ""),
                rec.get("fecha_recepcion", ""),
                rec.get("nombre_cliente", ""),
                str(sust.get("codigo", "") if sust else ""),
                str(sust.get("nombre", "") if sust else ""),
                rec.get("lote", ""),
                rec.get("cantidad", ""),
                self._nombre_usuario(rec.get("id_usuario_reviso")),
                self._nombre_usuario(rec.get("id_usuario_verifico")),
                rec.get("observaciones", ""),
            ]

            alt = PatternFill("solid", fgColor="F0FAF3") if row_idx % 2 == 0 else None

            for col_idx, val in enumerate(main_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align
                cell.border = thin
                if alt:
                    cell.fill = alt

            offset = len(main_headers)
            for ci, (field, _) in enumerate(CLIENTE_NUEVAS_FIELDS + CLIENTE_DEST_FIELDS):
                resp = str(rec.get(field, "") or "").strip().upper()
                cell = ws.cell(row=row_idx, column=offset + ci + 1, value=resp)
                cell.alignment = center_align
                cell.border = thin
                if resp == "SI":
                    cell.fill = si_fill
                elif resp == "NO":
                    cell.fill = no_fill
                elif resp in ("NA", "N/A", "NONE", ""):
                    cell.fill = na_fill

        for col_idx, w in enumerate(main_widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
        n_checks = len(CLIENTE_NUEVAS_FIELDS) + len(CLIENTE_DEST_FIELDS)
        for ci in range(n_checks):
            ws.column_dimensions[
                ws.cell(row=1, column=len(main_headers) + ci + 1).column_letter
            ].width = 20

        ws.row_dimensions[1].height = 36
        ws.freeze_panes = "A2"
        wb.save(filepath)

    # ------------------------------------------------------------------
    # Helpers de widget
    # ------------------------------------------------------------------

    def _btn(self, parent, text, bg, cmd):
        return tk.Button(
            parent, text=text, bg=bg, fg="white",
            font=("Segoe UI", 9, "bold"), relief="flat",
            bd=0, padx=10, pady=5, cursor="hand2", command=cmd,
        )

    def _btn_small(self, parent, text, bg, cmd):
        return tk.Button(
            parent, text=text, bg=bg, fg="white",
            font=("Segoe UI", 8, "bold"), relief="flat",
            bd=0, padx=5, pady=2, cursor="hand2", command=cmd,
        )


# ── Helpers de módulo ────────────────────────────────────────────────────────

def _load_all_sustancias() -> list:
    """Carga todas las sustancias (habilitadas e inhabilitadas) para resolución de nombres."""
    from database import get_db
    db = get_db()
    try:
        return db.get_sustancias()
    finally:
        db.close()


def _load_usuarios_into(dest: dict) -> None:
    """Carga usuarios en un dict {id: record}."""
    from logica import usuarios_logica as usr
    for u in usr.cargar():
        dest[int(u["id"])] = u
