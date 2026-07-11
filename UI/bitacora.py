import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, ttk

from config.config import COLORS
from logica import bitacora_logica as bit
from UI._mov_utils import apply_default_window, attach_treeview_sorting, draw_title, get_date_value, make_date_input
from UI._searchable_treeview import SearchableTreeview

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    _OPENPYXL = True
except ImportError:
    _OPENPYXL = False


def open_window(master):
    BitacoraWindow(master)


class BitacoraWindow(tk.Toplevel):
    def __init__(self, master):
        super().__init__(master)
        self.title("Sistema de Gestion - Bitacora")
        self.configure(bg=COLORS["secondary"])
        apply_default_window(self)

        self.v_modulo = tk.StringVar()
        self.v_tipo_operacion = tk.StringVar()
        self.v_usuario = tk.StringVar()
        self.v_id_registro = tk.StringVar()
        self._desde_habilitada = False
        self._hasta_habilitada = False

        # Pagination state (server-side)
        self._filters = {
            "desde": "",
            "hasta": "",
            "usuario": "",
            "tipo_operacion": "",
            "id_registro": "",
        }
        self._current_rows = []
        self._current_page = 1
        self._page_size = 25
        self._total_pages = 1
        self._total_records = 0

        draw_title(self, "Sistema de Gestion - Bitacora")
        self._build_ui()
        self._load_default()

    @staticmethod
    def _split_tipo_operacion(tipo_operacion):
        text = str(tipo_operacion or "").strip().upper()
        if not text:
            return "", ""
        if "-" in text:
            op, accion = text.split("-", 1)
            return op, accion
        return text, ""

    def _build_ui(self):
        wrap = tk.LabelFrame(
            self,
            text="  Auditoria de movimientos  ",
            bg=COLORS["secondary"],
            fg=COLORS["primary_dark"],
            font=("Segoe UI", 10, "bold"),
        )
        wrap.pack(fill="both", expand=True, padx=10, pady=10)

        top = tk.Frame(wrap, bg=COLORS["secondary"])
        top.pack(fill="x", padx=8, pady=8)

        frame_desde = tk.Frame(top, bg=COLORS["secondary"])
        frame_desde.pack(side="left", padx=(0, 8))
        self.w_desde = make_date_input(frame_desde, 0, 0, "Desde")
        self.w_desde.bind("<<DateEntrySelected>>", lambda _e: self._toggle_date_filter("desde", True))
        self._button(frame_desde, "Limpiar", "#B0B0B0", lambda: self._clear_date("desde")).grid(row=1, column=1, padx=(0, 8), pady=(0, 8))

        frame_hasta = tk.Frame(top, bg=COLORS["secondary"])
        frame_hasta.pack(side="left", padx=(0, 8))
        self.w_hasta = make_date_input(frame_hasta, 0, 0, "Hasta")
        self.w_hasta.bind("<<DateEntrySelected>>", lambda _e: self._toggle_date_filter("hasta", True))
        self._button(frame_hasta, "Limpiar", "#B0B0B0", lambda: self._clear_date("hasta")).grid(row=1, column=1, padx=(0, 8), pady=(0, 8))

        tk.Label(top, text="Usuario:", bg=COLORS["secondary"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
        tk.Entry(top, textvariable=self.v_usuario, width=10).pack(side="left", padx=(0, 8))
        tk.Label(top, text="Tipo Operacion:", bg=COLORS["secondary"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
        tk.Entry(top, textvariable=self.v_tipo_operacion, width=14).pack(side="left", padx=(0, 8))
        tk.Label(top, text="ID Registro:", bg=COLORS["secondary"], font=("Segoe UI", 9)).pack(side="left", padx=(0, 4))
        tk.Entry(top, textvariable=self.v_id_registro, width=8).pack(side="left", padx=(0, 8))

        self._button(top, "Filtrar", COLORS["primary"], self._apply_filters).pack(side="left", padx=(0, 6))
        self._button(top, "Borrar filtros", "#B0B0B0", self._clear_filters).pack(side="left")

        # Table in a frame with proper scrollbar
        tbl_frame = tk.Frame(wrap, bg=COLORS["secondary"])
        tbl_frame.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        tbl_frame.grid_rowconfigure(0, weight=1)
        tbl_frame.grid_columnconfigure(0, weight=1)

        cols = ("fecha_hora", "usuario", "tipo_operacion", "hoja", "id_registro", "detalle")
        self.tree = SearchableTreeview(
            tbl_frame, columns=cols,
            search_columns=["usuario", "tipo_operacion", "hoja", "detalle"],
        )
        self.tree.pack(fill="both", expand=True)

        for key, title, width in [
            ("fecha_hora", "Fecha y Hora", 145),
            ("usuario", "Usuario", 100),
            ("tipo_operacion", "Tipo Operacion", 110),
            ("hoja", "Hoja", 90),
            ("id_registro", "ID Registro", 80),
            ("detalle", "Detalle del Cambio", 560),
        ]:
            self.tree.heading(key, text=title)
            self.tree.column(key, width=width, anchor="center")
        attach_treeview_sorting(self.tree.tree)

        # Pagination controls
        pag_frame = tk.Frame(wrap, bg=COLORS["secondary"])
        pag_frame.pack(fill="x", padx=8, pady=(0, 4))

        btn_style = dict(bg=COLORS["primary"], fg="white", font=("Segoe UI", 9, "bold"),
                         relief="flat", bd=0, padx=8, pady=3, cursor="hand2")
        self.btn_pag_prev = tk.Button(pag_frame, text="< Anterior", command=self._prev_page, **btn_style)
        self.btn_pag_prev.pack(side="left", padx=2)
        self.lbl_page = tk.Label(pag_frame, text="Pagina 1 de 1", bg=COLORS["secondary"],
                                 fg=COLORS["text_dark"], font=("Segoe UI", 9))
        self.lbl_page.pack(side="left", padx=10)
        self.btn_pag_next = tk.Button(pag_frame, text="Siguiente >", command=self._next_page, **btn_style)
        self.btn_pag_next.pack(side="left", padx=2)
        tk.Label(pag_frame, text="Mostrar:", bg=COLORS["secondary"], fg=COLORS["text_dark"],
                 font=("Segoe UI", 9)).pack(side="left", padx=(15, 4))
        self.combo_page_size = ttk.Combobox(pag_frame, values=[25, 50, 100, 200],
                                            state="readonly", width=6, font=("Segoe UI", 9))
        self.combo_page_size.set("25")
        self.combo_page_size.bind("<<ComboboxSelected>>", self._on_page_size_change)
        self.combo_page_size.pack(side="left", padx=2)
        self.lbl_total_bit = tk.Label(pag_frame, text="", bg=COLORS["secondary"],
                                      fg=COLORS["text_muted"], font=("Segoe UI", 9, "italic"))
        self.lbl_total_bit.pack(side="left", padx=(12, 0))

        bottom = tk.Frame(self, bg=COLORS["secondary"])
        bottom.pack(fill="x", padx=10, pady=(0, 10))
        self._button(bottom, "Actualizar", COLORS["primary"], self._load_default).pack(side="right", padx=(6, 0))
        self._button(bottom, "Salir", "#6C757D", self.destroy).pack(side="right")
        self._button(bottom, "📥 Exportar Excel", "#1565C0", self._export_excel).pack(side="left")

    def _button(self, parent, text, bg, cmd):
        return tk.Button(parent, text=text, bg=bg, fg="white", font=("Segoe UI", 9, "bold"),
                         relief="flat", bd=0, padx=10, pady=5, cursor="hand2", command=cmd)

    def _toggle_date_filter(self, which, enabled):
        if which == "desde":
            self._desde_habilitada = enabled
        else:
            self._hasta_habilitada = enabled

    def _clear_date(self, which):
        self._toggle_date_filter(which, False)
        widget = self.w_desde if which == "desde" else self.w_hasta
        var = getattr(widget, "_fallback_var", None)
        if var is not None:
            var.set("")
            return
        try:
            widget.delete(0, "end")
        except Exception:
            pass

    # ------------------------------------------------------------------
    # Pagination
    # ------------------------------------------------------------------
    def _load_default(self):
        self._filters = {"desde": "", "hasta": "", "usuario": "", "tipo_operacion": "", "id_registro": ""}
        self._current_page = 1
        self._total_records = bit.contar_filtrado(**self._filters)
        self._total_pages = max(1, -(-self._total_records // self._page_size))
        self._render_page()

    def _apply_filters(self):
        self._filters = {
            "desde": get_date_value(self.w_desde) if self._desde_habilitada else "",
            "hasta": get_date_value(self.w_hasta) if self._hasta_habilitada else "",
            "usuario": self.v_usuario.get().strip(),
            "tipo_operacion": self.v_tipo_operacion.get().strip(),
            "id_registro": self.v_id_registro.get().strip(),
        }
        self._current_page = 1
        self._total_records = bit.contar_filtrado(**self._filters)
        self._total_pages = max(1, -(-self._total_records // self._page_size))
        self._render_page()

    def _render_page(self):
        self.tree.clear()
        start = (self._current_page - 1) * self._page_size
        self._current_rows = bit.filtrar_paginado(
            **self._filters,
            limit=self._page_size,
            offset=start,
        )
        for r in self._current_rows:
            hoja, accion = self._split_tipo_operacion(r.get("tipo_operacion", ""))
            self.tree.insert("", "end", values=(
                r.get("fecha_hora", ""),
                r.get("usuario", ""),
                accion,
                hoja,
                r.get("id_registro", ""),
                self._build_detalle(r),
            ))
        self._update_pagination_buttons()

    @staticmethod
    def _build_detalle(row):
        campo = str(row.get("campo", "") or "").strip()
        anterior = str(row.get("valor_anterior", "") or "").strip()
        nuevo = str(row.get("valor_nuevo", "") or "").strip()
        tipo_op = str(row.get("tipo_operacion", "") or "").strip().upper()

        accion = "operacion"
        if "-" in tipo_op:
            _mod, _acc = tipo_op.split("-", 1)
            accion = _acc.lower()
        elif tipo_op:
            accion = tipo_op.lower()

        if campo in {"DETALLE", "REGISTRO"}:
            if nuevo:
                return (
                    f"Se registro la accion de {accion} con el siguiente detalle: {nuevo}"
                )
            if anterior:
                return (
                    f"Se registro la accion de {accion} con el siguiente detalle: {anterior}"
                )
            return f"Se registro la accion de {accion}, pero no se encontraron datos adicionales del cambio."

        if anterior and nuevo:
            return (
                f"En la accion de {accion}, el campo '{campo}' fue modificado. "
                f"Valor anterior: '{anterior}'. Valor nuevo: '{nuevo}'."
            )
        if nuevo:
            return (
                f"En la accion de {accion}, se establecio el campo '{campo}' "
                f"con el valor: '{nuevo}'."
            )
        if anterior:
            return (
                f"En la accion de {accion}, el campo '{campo}' conserva el valor previo: '{anterior}'."
            )
        return f"Se registro la accion de {accion}, pero el sistema no reporto cambios especificos en los campos auditados."

    def _prev_page(self):
        if self._current_page > 1:
            self._current_page -= 1
            self._render_page()

    def _next_page(self):
        if self._current_page < self._total_pages:
            self._current_page += 1
            self._render_page()

    def _on_page_size_change(self, _event=None):
        try:
            self._page_size = int(self.combo_page_size.get())
            self._current_page = 1
            self._total_pages = max(1, -(-self._total_records // self._page_size))
            self._render_page()
        except ValueError:
            pass

    def _update_pagination_buttons(self):
        if not hasattr(self, "btn_pag_prev"):
            return
        self.btn_pag_prev.config(state=tk.NORMAL if self._current_page > 1 else tk.DISABLED)
        self.btn_pag_next.config(state=tk.NORMAL if self._current_page < self._total_pages else tk.DISABLED)
        self.lbl_page.config(text=f"Pagina {self._current_page} de {max(self._total_pages, 1)}")
        if hasattr(self, "lbl_total_bit"):
            self.lbl_total_bit.config(text=f"({self._total_records} registros)")

    def _clear_filters(self):
        self._clear_date("desde")
        self._clear_date("hasta")
        self.v_usuario.set("")
        self.v_tipo_operacion.set("")
        self.v_id_registro.set("")
        self._load_default()

    # ------------------------------------------------------------------
    # Exportación a Excel
    # ------------------------------------------------------------------

    def _export_excel(self):
        if not _OPENPYXL:
            messagebox.showerror(
                "Dependencia faltante",
                "openpyxl no está instalado.\nInstala con: pip install openpyxl",
                parent=self,
            )
            return

        rows = bit.filtrar_paginado(**self._filters, limit=max(1, self._total_records), offset=0)
        if not rows:
            messagebox.showwarning("Sin datos", "No hay registros para exportar.", parent=self)
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"bitacora_{timestamp}.xlsx"
        filepath = filedialog.asksaveasfilename(
            parent=self,
            title="Guardar Bitácora",
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
        )
        if not filepath:
            return

        try:
            self._write_excel(filepath, rows)
            messagebox.showinfo("Exportado",
                                f"Bitácora exportada exitosamente:\n{filepath}", parent=self)
        except PermissionError:
            messagebox.showerror(
                "Error",
                "No se pudo guardar el archivo. Verifica que no esté abierto en Excel.",
                parent=self,
            )
        except Exception as exc:
            messagebox.showerror("Error", f"Error al exportar:\n{exc}", parent=self)

    @staticmethod
    def _write_excel(filepath: str, rows: list) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Bitácora"

        header_font  = Font(bold=True, color="FFFFFF", size=10)
        header_fill  = PatternFill("solid", fgColor="0F4C97")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        headers = [
            ("fecha_hora",      "Fecha y Hora",          18),
            ("usuario",         "Usuario",                14),
            ("tipo_operacion",  "Módulo",                 14),
            ("accion",          "Acción",                 14),
            ("id_registro",     "ID Registro",            11),
            ("detalle",         "Detalle del Cambio",     60),
        ]

        for col_idx, (_, title, width) in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font  = header_font
            cell.fill  = header_fill
            cell.alignment = header_align
            cell.border = thin
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = 28

        def _split(tipo_op):
            text = str(tipo_op or "").strip().upper()
            if "-" in text:
                mod, acc = text.split("-", 1)
                return mod, acc
            return text, ""

        def _build_detalle(r):
            campo = str(r.get("campo", "") or "").strip()
            anterior = str(r.get("valor_anterior", "") or "").strip()
            nuevo = str(r.get("valor_nuevo", "") or "").strip()
            tipo_op = str(r.get("tipo_operacion", "") or "").strip().upper()
            accion = "operacion"
            if "-" in tipo_op:
                _, _acc = tipo_op.split("-", 1)
                accion = _acc.lower()
            elif tipo_op:
                accion = tipo_op.lower()
            if campo in {"DETALLE", "REGISTRO"}:
                return (nuevo or anterior or f"Se registró la acción de {accion}.")
            if anterior and nuevo:
                return f"Campo '{campo}': '{anterior}' → '{nuevo}'"
            if nuevo:
                return f"Campo '{campo}' establecido: '{nuevo}'"
            if anterior:
                return f"Campo '{campo}' previo: '{anterior}'"
            return f"Acción: {accion}"

        alt_fill = PatternFill("solid", fgColor="F0F4FC")

        for row_idx, r in enumerate(rows, start=2):
            mod, acc = _split(r.get("tipo_operacion", ""))
            values = [
                r.get("fecha_hora", ""),
                r.get("usuario", ""),
                mod,
                acc,
                r.get("id_registro", ""),
                _build_detalle(r),
            ]
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = center_align if col_idx < 6 else Alignment(wrap_text=True)
                cell.border = thin
                if fill:
                    cell.fill = fill

        ws.freeze_panes = "A2"
        wb.save(filepath)